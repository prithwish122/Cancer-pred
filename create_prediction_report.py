"""
Script to create a prediction report Excel file with:
- Best Model Used (ranked by Recall, then AUC)
- Performance Metrics
- Individual Predictions with Patient IDs

Uses the same optimized pipeline as XAI_ByteA_FeatureEngineeringModelTraining.py:
- PCA for dimensionality reduction
- GridSearchCV with recall-focused optimization
- Proper train/test split with no data leakage
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference

# Import the training pipeline components
import sys
import warnings
warnings.filterwarnings("ignore")

from sklearn.model_selection import train_test_split, StratifiedKFold, GridSearchCV
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.impute import KNNImputer
from sklearn.feature_selection import VarianceThreshold
from sklearn.decomposition import PCA
from sklearn.metrics import (
    accuracy_score,
    precision_score,
    recall_score,
    f1_score,
    roc_auc_score,
    confusion_matrix,
    roc_curve,
)
from sklearn.ensemble import RandomForestClassifier, GradientBoostingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.svm import SVC
from sklearn.neighbors import KNeighborsClassifier
from sklearn.naive_bayes import GaussianNB
from sklearn.tree import DecisionTreeClassifier

# Try to import XGBoost
try:
    from xgboost import XGBClassifier
    XGBOOST_AVAILABLE = True
except ImportError:
    XGBClassifier = None
    XGBOOST_AVAILABLE = False

# Try to import SMOTE, but make it optional
try:
    from imblearn.over_sampling import SMOTE
    SMOTE_AVAILABLE = True
except (ImportError, TypeError):
    SMOTE_AVAILABLE = False
    print("Warning: SMOTE not available, will use class_weight balancing instead")

def create_prediction_report():
    """
    Creates an Excel file with model performance metrics and predictions
    Uses the same optimized pipeline as the training script: PCA, GridSearchCV, recall-focused
    """
    FAST_MODE = True  # Match training script setting
    
    print("Loading dataset...")
    df = pd.read_excel('AI_ByteA_CleanedDataset.xlsx', engine='openpyxl')
    print(f"Raw data shape: {df.shape}")
    
    # Store PatientID if available
    patient_ids = None
    if 'PatientID' in df.columns:
        patient_ids = df['PatientID'].copy()
    
    # ---------- Basic cleaning: drop duplicates ----------
    print(f"Number of duplicate rows before drop: {df.duplicated().sum()}")
    df = df.drop_duplicates().reset_index(drop=True)
    print(f"Data shape after dropping duplicates: {df.shape}")
    
    # ---------- Coerce numeric columns & cap extreme outliers ----------
    num_cols_raw = df.select_dtypes(include=[np.number]).columns.tolist()
    for col in num_cols_raw:
        # Coerce any bad formats to NaN to avoid silent string issues
        df[col] = pd.to_numeric(df[col], errors='coerce')
    
    numeric_df = df[num_cols_raw]
    q1 = numeric_df.quantile(0.01)
    q99 = numeric_df.quantile(0.99)
    # Winsorize numeric features so extreme outliers don't distort scaling
    df[num_cols_raw] = numeric_df.clip(lower=q1, upper=q99, axis=1)
    
    # ---------- Simple feature engineering (safe, recall-friendly) ----------
    # Example: derive AgeBucket if Age exists to let models capture non-linear age-risk patterns
    if 'Age' in df.columns:
        print("Creating AgeBucket feature from Age...")
        df['AgeBucket'] = pd.cut(
            df['Age'],
            bins=[0, 30, 45, 60, 75, 120],
            labels=[0, 1, 2, 3, 4],
            include_lowest=True
        ).astype('int64')
    
    # ---------- Encode categoricals ----------
    print("Encoding categoricals...")
    label_encoders = {}
    categorical_cols = df.select_dtypes(include=['object']).columns
    for col in categorical_cols:
        if col != 'PatientID':
            le = LabelEncoder()
            df[col] = le.fit_transform(df[col].astype(str))
            label_encoders[col] = le
    
    # ---------- Target & features ----------
    target_col = 'Diagnosis' if 'Diagnosis' in df.columns else df.columns[-1]
    drop_cols = ['PatientID', target_col] if 'PatientID' in df.columns else [target_col]
    
    X = df.drop(columns=drop_cols, errors='ignore')
    y = df[target_col].astype(int)
    
    # ---------- Train / test split (before fitting imputers/scalers to avoid leakage) ----------
    test_size = 0.2 if FAST_MODE else 0.1
    print(f"Train/test split (test_size={test_size})...")
    X_train, X_test, y_train, y_test = train_test_split(
        X, y, test_size=test_size, random_state=42, stratify=y
    )
    print(f"Train: {X_train.shape}, Test: {X_test.shape}")
    
    # ---------- Impute (fit only on train, then transform test) ----------
    print("Imputing missing values on train with KNNImputer (no leakage)...")
    imputer = KNNImputer(n_neighbors=3 if FAST_MODE else 5)
    X_train_imp = pd.DataFrame(imputer.fit_transform(X_train), columns=X_train.columns)
    X_test_imp = pd.DataFrame(imputer.transform(X_test), columns=X_test.columns)
    
    # ---------- Remove very low variance (fit on train only) ----------
    print("Variance thresholding (fit on train only)...")
    var_threshold = VarianceThreshold(threshold=0.0 if FAST_MODE else 0.0001)
    X_train_var = var_threshold.fit_transform(X_train_imp)
    X_test_var = var_threshold.transform(X_test_imp)
    
    # ---------- Scale (fit scaler on train only to avoid test leakage) ----------
    print("Scaling features with StandardScaler (fit on train only)...")
    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train_var)
    X_test_scaled = scaler.transform(X_test_var)
    
    # ---------- PCA for dimensionality reduction (helps prevent overfitting) ----------
    print("Applying PCA for dimensionality reduction (preserve ~95% variance)...")
    # n_components=0.95 chooses the number of components that explain ~95% of variance
    pca = PCA(n_components=0.95, random_state=42)
    X_train_pca = pca.fit_transform(X_train_scaled)
    X_test_pca = pca.transform(X_test_scaled)
    print(f"PCA-transformed shape: train={X_train_pca.shape}, test={X_test_pca.shape}")
    
    # Downstream models will now use the PCA feature space
    X_train, X_test = X_train_pca, X_test_pca
    
    # ---------- Balance (use SMOTE if available) ----------
    if SMOTE_AVAILABLE:
        try:
            print("Balancing with SMOTE...")
            smote = SMOTE(random_state=42, k_neighbors=3 if FAST_MODE else 5)
            X_train_bal, y_train_bal = smote.fit_resample(X_train, y_train)
            print(f"Balanced train: {X_train_bal.shape}, class ratio: {np.bincount(y_train_bal)}")
        except Exception as e:
            print(f"SMOTE failed ({e}), using original data with class_weight balancing")
            X_train_bal, y_train_bal = X_train, y_train
    else:
        print("SMOTE not available – proceeding without oversampling (using original train set).")
        X_train_bal, y_train_bal = X_train, y_train
    
    # Prepare full dataset for predictions (apply same preprocessing pipeline)
    print("Preparing full dataset for predictions...")
    # Apply the same preprocessing steps to full dataset (fit on train, transform full)
    X_full_imp = pd.DataFrame(imputer.transform(X), columns=X.columns)
    X_full_var = var_threshold.transform(X_full_imp)
    X_full_scaled = scaler.transform(X_full_var)
    X_full_pca = pca.transform(X_full_scaled)
    X_full = X_full_pca
    y_full = y.values
    
    # Get all PatientIDs for full dataset
    if patient_ids is not None:
        patient_ids_full = patient_ids.values if hasattr(patient_ids, 'values') else list(patient_ids)
    else:
        # Generate PatientIDs if not available
        patient_ids_full = [f"PID{i+1}" for i in range(len(y_full))]
    
    print(f"Full dataset for predictions: {X_full.shape}")
    
    # ---------- Models (base configs) ----------
    print("Building base models...")
    
    models = {
        'RandomForest': RandomForestClassifier(
            n_estimators=400 if FAST_MODE else 900,
            max_depth=None,
            min_samples_split=4,
            min_samples_leaf=2,
            max_features='sqrt',
            bootstrap=True,
            criterion='gini',
            class_weight='balanced',  # balanced for recall without hardcoding thresholds
            random_state=42,
            n_jobs=-1,
            verbose=0
        ),
        'GradientBoosting': GradientBoostingClassifier(
            n_estimators=300 if FAST_MODE else 700,
            learning_rate=0.05 if FAST_MODE else 0.02,
            max_depth=2,
            subsample=0.9,
            random_state=42
        ),
        'LogisticRegression': LogisticRegression(
            max_iter=3000 if FAST_MODE else 6000,
            C=1.0 if FAST_MODE else 3.0,
            penalty='l2',
            solver='lbfgs',
            class_weight='balanced',
            n_jobs=None,
            verbose=0
        ),
        'SVM': SVC(
            probability=True,
            kernel='rbf',
            C=1.0 if FAST_MODE else 3.0,
            gamma='scale',
            class_weight='balanced',
            cache_size=1024,
            tol=1e-3 if FAST_MODE else 1e-4,
            random_state=42
        ),
        'KNN': KNeighborsClassifier(
            n_neighbors=7,
            weights='distance',
            algorithm='auto',
            p=2
        ),
        'NaiveBayes': GaussianNB(),
        'DecisionTree': DecisionTreeClassifier(
            max_depth=None,
            min_samples_split=4,
            min_samples_leaf=2,
            random_state=42
        ),
    }
    
    if XGBOOST_AVAILABLE:
        models['XGBoost'] = XGBClassifier(
            n_estimators=300 if FAST_MODE else 600,
            max_depth=3 if FAST_MODE else 5,
            learning_rate=0.05 if FAST_MODE else 0.03,
            subsample=0.9,
            colsample_bytree=0.8,
            objective='binary:logistic',
            eval_metric='logloss',
            n_jobs=-1,
            random_state=42,
            tree_method='hist'
        )
        print("XGBoost added to model list.")
    else:
        print("XGBoost not available; skipping XGBoost model.")
    
    # ---------- Hyperparameter tuning with GridSearchCV (RandomForest, GradientBoosting, XGBoost) ----------
    print("Preparing hyperparameter search with GridSearchCV (recall-focused)...")
    cv_splits = 3 if FAST_MODE else 5
    cv = StratifiedKFold(n_splits=cv_splits, shuffle=True, random_state=42)
    
    param_grids = {
        'RandomForest': {
            'n_estimators': [300, 400, 600, 900],
            'max_depth': [None, 10, 20, 40],
            'min_samples_split': [2, 4, 6],
            'min_samples_leaf': [1, 2, 3],
            'max_features': ['sqrt', 'log2'],
        },
        'GradientBoosting': {
            'n_estimators': [200, 300, 500, 700],
            'learning_rate': [0.01, 0.03, 0.05],
            'max_depth': [2, 3, 4],
            'subsample': [0.8, 0.9, 1.0],
        },
    }
    
    if XGBOOST_AVAILABLE:
        param_grids['XGBoost'] = {
            'n_estimators': [200, 300, 500],
            'max_depth': [3, 4, 5],
            'learning_rate': [0.01, 0.03, 0.05],
            'subsample': [0.8, 0.9, 1.0],
            'colsample_bytree': [0.7, 0.8],
        }
    
    for m_name, grid in param_grids.items():
        if m_name not in models:
            continue
        base_model = models[m_name]
        print(f"Tuning {m_name} with GridSearchCV (cv={cv_splits})...")
        search = GridSearchCV(
            estimator=base_model,
            param_grid=grid,
            scoring={
                "accuracy": "accuracy",
                "precision": "precision",
                "recall": "recall",
                "f1": "f1",
                "roc_auc": "roc_auc",
            },
            # Choose the configuration with the highest cross-validated recall
            refit="recall",
            n_jobs=-1,
            cv=cv,
            verbose=1 if not FAST_MODE else 0
        )
        search.fit(X_train_bal, y_train_bal)
        models[m_name] = search.best_estimator_
        print(f"{m_name} best params: {search.best_params_}")
    
    # ---------- Train & evaluate ----------
    print("Training tuned models & evaluating...")
    results = []
    trained_models = {}
    
    for name, model in models.items():
        print(f"Training {name}...")
        model.fit(X_train_bal, y_train_bal)
        trained_models[name] = model
        
        # Base predictions at model's default threshold
        y_pred = model.predict(X_test)
        
        # Probabilities if available; else use decision_function or skip AUC/ROC
        if hasattr(model, "predict_proba"):
            y_proba = model.predict_proba(X_test)[:, 1]
        elif hasattr(model, "decision_function"):
            df_scores = model.decision_function(X_test)
            # scale to 0-1 for auc/roc
            y_proba = (df_scores - df_scores.min()) / (df_scores.max() - df_scores.min() + 1e-12)
        else:
            y_proba = None
        
        accuracy = accuracy_score(y_test, y_pred)
        precision = precision_score(y_test, y_pred, zero_division=0)
        recall = recall_score(y_test, y_pred, zero_division=0)
        f1 = f1_score(y_test, y_pred, zero_division=0)
        auc_score = roc_auc_score(y_test, y_proba) if y_proba is not None else 0.0
        
        results.append({
            'Model': name,
            'Accuracy': accuracy,
            'Precision': precision,
            'Recall': recall,
            'F1': f1,
            'AUC': auc_score
        })
        print(f"{name} - Recall: {recall:.4f}, AUC: {auc_score:.4f}")
    
    results_df = pd.DataFrame(results)
    
    # Sort models by Recall (primary) and AUC (secondary) so ranking reflects real test performance
    results_df = results_df.sort_values(['Recall', 'AUC'], ascending=[False, False]).reset_index(drop=True)
    results_df['Rank'] = results_df.index + 1
    
    print("\nModel Performance Summary (ranked by Recall, then AUC):")
    print(results_df[['Rank', 'Model', 'Accuracy', 'F1', 'Precision', 'Recall', 'AUC']].to_string(index=False))
    
    # Select best model by highest Recall; if tie, highest AUC wins
    best_model_name = results_df.iloc[0]['Model']
    print(f"\nBest model by Recall (tie-broken by AUC): {best_model_name}")
    
    best_model = trained_models[best_model_name]
    best_metrics = results_df[results_df['Model'] == best_model_name].iloc[0]
    
    # Display all results
    print(f"\nBest Model: {best_model_name}")
    print(f"Best Model Recall: {best_metrics['Recall']:.4f} ({best_metrics['Recall']*100:.2f}%)")
    print(f"Best Model Accuracy: {best_metrics['Accuracy']:.4f} ({best_metrics['Accuracy']*100:.2f}%)")
    
    # Use the ALREADY TRAINED best model (on train split) for predictions.
    # This avoids overfitting by not refitting on the full dataset.
    best_model_final = trained_models[best_model_name]
    
    # Get predictions from best model on FULL DATASET
    print("Generating predictions for full dataset with trained best model (no refit)...")
    y_pred_full = best_model_final.predict(X_full)
    if hasattr(best_model_final, "predict_proba"):
        y_proba_full = best_model_final.predict_proba(X_full)[:, 1]
    else:
        y_proba_full = y_pred_full.astype(float)
    
    # Create predictions dataframe for full dataset
    predictions_df = pd.DataFrame({
        'PatientID': patient_ids_full,
        'Actual': y_full,
        'Predicted': y_pred_full,
        'Probability': y_proba_full,
        'Correct': (y_full == y_pred_full).astype(int)
    })
    
    # Calculate final metrics
    final_accuracy = predictions_df['Correct'].mean()
    final_precision = precision_score(y_full, y_pred_full, zero_division=0)
    final_recall = recall_score(y_full, y_pred_full, zero_division=0)
    final_f1 = f1_score(y_full, y_pred_full, zero_division=0)
    
    if hasattr(best_model_final, "predict_proba"):
        y_proba_full_calc = best_model_final.predict_proba(X_full)[:, 1]
        final_auc = roc_auc_score(y_full, y_proba_full_calc)
    else:
        y_proba_full_calc = y_proba_full
        final_auc = 0.0
    
    print(f"\nFinal Model Metrics on Full Dataset:")
    print(f"  Accuracy: {final_accuracy:.4f} ({final_accuracy*100:.2f}%)")
    print(f"  Precision: {final_precision:.4f} ({final_precision*100:.2f}%)")
    print(f"  Recall: {final_recall:.4f} ({final_recall*100:.2f}%)")
    print(f"  F1: {final_f1:.4f} ({final_f1*100:.2f}%)")
    print(f"  AUC: {final_auc:.4f}")
    
    # Prepare additional diagnostics for Sheet 2 (ROC, Confusion Matrix, Feature Importance)
    # Confusion matrix on full dataset
    cm_full = confusion_matrix(y_full, y_pred_full)
    
    # ROC curve data (only if probabilities are meaningful)
    fpr_full, tpr_full, _ = roc_curve(y_full, y_proba_full_calc if 'y_proba_full_calc' in locals() else y_proba_full)
    
    # Feature importances (if available)
    feature_importances = None
    if hasattr(best_model_final, "feature_importances_"):
        feature_importances = best_model_final.feature_importances_
    
    # Create Excel file
    print("Creating Excel report...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Prediction_Report"
    
    # Define styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font = Font(bold=True, size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Section 1: Best Model Used
    ws['A1'] = "Best Model Used"
    ws['A1'].font = title_font
    ws['B1'] = best_model_name
    ws['B1'].font = Font(bold=True, size=11)
    
    # Section 2: Performance Metrics
    ws['A3'] = "Performance Metrics"
    ws.merge_cells('A3:G3')
    ws['A3'].font = title_font
    ws['A3'].alignment = center_align
    
    # Performance Metrics Headers
    headers = ['Model', '', 'Accuracy', 'Precision', 'Recall', 'F1', 'AUC']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Performance Metrics Data (only best model)
    row = 5
    try:
        ws.cell(row=row, column=1).value = best_model_name
        ws.cell(row=row, column=3).value = round(float(best_metrics.get('Accuracy', 0)), 3)
        ws.cell(row=row, column=4).value = round(float(best_metrics.get('Precision', 0)), 6)
        ws.cell(row=row, column=5).value = round(float(best_metrics.get('Recall', 0)), 3)
        ws.cell(row=row, column=6).value = round(float(best_metrics.get('F1', 0)), 6)
        ws.cell(row=row, column=7).value = round(float(best_metrics.get('AUC', 0)), 5)
    except Exception as e:
        print(f"Warning: Error writing metrics to Excel: {e}")
        # Use default values
        ws.cell(row=row, column=1).value = best_model_name
        ws.cell(row=row, column=3).value = final_accuracy
        ws.cell(row=row, column=4).value = 0.0
        ws.cell(row=row, column=5).value = 0.0
        ws.cell(row=row, column=6).value = 0.0
        ws.cell(row=row, column=7).value = 0.0
    
    # Apply formatting to data row
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        if col == 1:
            cell.alignment = Alignment(horizontal='left')
        else:
            cell.alignment = center_align
    
    # Section 3: Predictions
    ws['A7'] = "Predictions"
    ws['A7'].font = title_font
    
    # Predictions Headers
    pred_headers = ['PatientID', '', 'Actual', 'Predicted', 'Probability', 'Correct']
    for col_idx, header in enumerate(pred_headers, start=1):
        cell = ws.cell(row=8, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Predictions Data
    for idx, (_, row_data) in enumerate(predictions_df.iterrows(), start=9):
        ws.cell(row=idx, column=1).value = str(row_data['PatientID'])
        ws.cell(row=idx, column=3).value = int(row_data['Actual'])
        ws.cell(row=idx, column=4).value = int(row_data['Predicted'])
        ws.cell(row=idx, column=5).value = round(row_data['Probability'], 6)
        ws.cell(row=idx, column=6).value = int(row_data['Correct'])
        
        # Apply formatting
        for col in range(1, 7):
            cell = ws.cell(row=idx, column=col)
            cell.border = border
            if col == 1:
                cell.alignment = Alignment(horizontal='left')
            else:
                cell.alignment = center_align
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 2
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    
    # -------- Sheet 2: ROC, Confusion Matrix, Feature Importance --------
    ws2 = wb.create_sheet(title="ROC_CM_Features")
    
    # ROC Curve Section (data + line chart)
    ws2['A1'] = "ROC Curve (Full Dataset)"
    ws2['A1'].font = title_font
    
    roc_headers = ['FPR', 'TPR']
    for col_idx, header in enumerate(roc_headers, start=1):
        cell = ws2.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    for i, (fpr_val, tpr_val) in enumerate(zip(fpr_full, tpr_full), start=3):
        ws2.cell(row=i, column=1).value = float(fpr_val)
        ws2.cell(row=i, column=2).value = float(tpr_val)
        for col in range(1, 3):
            cell = ws2.cell(row=i, column=col)
            cell.border = border
            cell.alignment = center_align
    
    # Create ROC line chart
    roc_chart = LineChart()
    roc_chart.title = "ROC Curve"
    roc_chart.y_axis.title = "TPR"
    roc_chart.x_axis.title = "FPR"
    data_ref = Reference(ws2, min_col=2, min_row=2, max_row=2 + len(tpr_full))
    cats_ref = Reference(ws2, min_col=1, min_row=3, max_row=2 + len(fpr_full))
    roc_chart.add_data(data_ref, titles_from_data=True)
    roc_chart.set_categories(cats_ref)
    roc_chart.height = 12
    roc_chart.width = 18
    ws2.add_chart(roc_chart, "E2")
    
    # Confusion Matrix Section (start after ROC data)
    cm_start_row = len(fpr_full) + 5
    ws2.cell(row=cm_start_row, column=1).value = "Confusion Matrix (Full Dataset)"
    ws2.cell(row=cm_start_row, column=1).font = title_font
    
    cm_headers = ['', 'Pred 0', 'Pred 1']
    for col_idx, header in enumerate(cm_headers, start=1):
        cell = ws2.cell(row=cm_start_row + 1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    # Rows for True 0 and True 1
    ws2.cell(row=cm_start_row + 2, column=1).value = 'True 0'
    ws2.cell(row=cm_start_row + 3, column=1).value = 'True 1'
    ws2.cell(row=cm_start_row + 2, column=2).value = int(cm_full[0, 0])
    ws2.cell(row=cm_start_row + 2, column=3).value = int(cm_full[0, 1])
    ws2.cell(row=cm_start_row + 3, column=2).value = int(cm_full[1, 0])
    ws2.cell(row=cm_start_row + 3, column=3).value = int(cm_full[1, 1])
    
    for r in range(cm_start_row + 2, cm_start_row + 4):
        for c in range(1, 4):
            cell = ws2.cell(row=r, column=c)
            cell.border = border
            cell.alignment = center_align
    
    # Confusion matrix bar chart
    cm_chart = BarChart()
    cm_chart.title = "Confusion Matrix Counts"
    cm_chart.y_axis.title = "Count"
    cm_chart.x_axis.title = "Class / Prediction"
    cm_data = Reference(ws2, min_col=2, min_row=cm_start_row + 1, max_col=3, max_row=cm_start_row + 3)
    cm_cats = Reference(ws2, min_col=1, min_row=cm_start_row + 2, max_row=cm_start_row + 3)
    cm_chart.add_data(cm_data, titles_from_data=True)
    cm_chart.set_categories(cm_cats)
    cm_chart.height = 12
    cm_chart.width = 18
    ws2.add_chart(cm_chart, f"E{cm_start_row}")
    
    # Feature Importance Section (if available)
    fi_start_row = cm_start_row + 6
    ws2.cell(row=fi_start_row, column=1).value = "Feature Importances"
    ws2.cell(row=fi_start_row, column=1).font = title_font
    
    fi_headers = ['Feature_Index', 'Importance']
    for col_idx, header in enumerate(fi_headers, start=1):
        cell = ws2.cell(row=fi_start_row + 1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    
    if feature_importances is not None:
        # Sort features by importance (descending) and keep top N for a cleaner chart
        top_n = min(20, len(feature_importances))
        sorted_idx = np.argsort(feature_importances)[-top_n:][::-1]
        for rank, idx in enumerate(sorted_idx):
            imp = feature_importances[idx]
            row_idx = fi_start_row + 2 + rank
            ws2.cell(row=row_idx, column=1).value = f"F_{idx}"
            ws2.cell(row=row_idx, column=2).value = float(imp)
            for c in range(1, 3):
                cell = ws2.cell(row=row_idx, column=c)
                cell.border = border
                cell.alignment = center_align
        
        # Feature importance bar chart (top N features)
        fi_last_row = fi_start_row + 1 + top_n
        fi_chart = BarChart()
        fi_chart.title = "Feature Importances"
        fi_chart.y_axis.title = "Importance"
        fi_chart.x_axis.title = "Feature"
        fi_data = Reference(ws2, min_col=2, min_row=fi_start_row + 1, max_row=fi_last_row)
        fi_cats = Reference(ws2, min_col=1, min_row=fi_start_row + 2, max_row=fi_last_row)
        fi_chart.add_data(fi_data, titles_from_data=True)
        fi_chart.set_categories(fi_cats)
        fi_chart.height = 12
        fi_chart.width = 18
        ws2.add_chart(fi_chart, f"E{fi_start_row}")
    else:
        ws2.cell(row=fi_start_row + 2, column=1).value = "No feature_importances_ attribute for this model."
    
    # Adjust some widths on sheet 2
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 12
    
    # Save file with error handling
    output_file = 'Prediction_Report.xlsx'
    try:
        wb.save(output_file)
    except PermissionError:
        # If file is open, try with a timestamp
        import datetime
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f'Prediction_Report_{timestamp}.xlsx'
        print(f"Original file is open, saving as: {output_file}")
        wb.save(output_file)
    except Exception as e:
        print(f"Error saving file: {e}")
        raise
    print(f"\n[SUCCESS] Prediction report saved to: {output_file}")
    print(f"[SUCCESS] Best Model: {best_model_name} (ranked by Recall)")
    print(f"[SUCCESS] Total Predictions: {len(predictions_df)}")
    print(f"[SUCCESS] Correct Predictions: {predictions_df['Correct'].sum()}")
    print(f"[SUCCESS] Accuracy: {predictions_df['Correct'].mean():.4f}")
    print(f"[SUCCESS] Recall: {final_recall:.4f}")
    
    return output_file

if __name__ == "__main__":
    create_prediction_report()

